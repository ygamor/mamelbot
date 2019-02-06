import discord
import asyncio
import openpyxl
import Weather
import random
import datetime
import Knoledge
import AI
import Dictionary
import enforce


client = discord.Client()

@client.event
async def on_ready():
    print('login')
    print(client.user.name)
    print(client.user.id)
    print('------------------------')


@client.event
async def on_message(message):
    file = openpyxl.load_workbook("쿨타임.xlsx")
    sheet = file.active
    file1 = openpyxl.load_workbook("밴.xlsx")
    sheet1 = file1.active
    kk = 0

    if "<@249706579912294400>" in message.content:
        file = openpyxl.load_workbook('상태.xlsx')
        sheet = file.active
        if sheet["A1"].value != "활동" and message.author.id != '536190598806503435' and message.author.id != '249706579912294400':
            await client.send_message(message.channel, "지금 관리자는 " + str(sheet["A1"].value) + "중입니다.\n+메시지를 통해 관리자방에 쪽지를 남겨둘수있어요.")

    if message.content.startswith("!") or message.content.startswith("+"):
        i = 1
        while True:
            if message.author.id == "249706579912294400":
                break
            if sheet["A" + str(i)].value == message.author.id:
                if int(sheet["B" + str(i)].value) > int(datetime.datetime.today().strftime("%Y%m%d%H%M%S")):
                    a = datetime.datetime.today() + datetime.timedelta(seconds=4)
                    sheet["B" + str(i)].value = a.strftime("%Y%m%d%H%M%S")
                    await client.send_message(message.channel, "<@" + str(message.author.id) + ">" + "명령어를 천천히 입력해주세요. 쿨타임: `4`초")
                    file.save("쿨타임.xlsx")
                else:
                    a = datetime.datetime.today() + datetime.timedelta(seconds=4)
                    sheet["B" + str(i)].value = a.strftime("%Y%m%d%H%M%S")
                    file.save("쿨타임.xlsx")
                break

            if sheet["A" + str(i)].value == None:
                a = datetime.datetime.today() + datetime.timedelta(seconds=4)
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = a.strftime("%Y%m%d%H%M%S")
                file.save("쿨타임.xlsx")
                break
            i = i + 1

        i = 1
        while True:
            if sheet1["A" + str(i)].value == message.author.id:
                msg = "차단된 회원입니다."
                embed = discord.Embed(description=msg, color=0xee2f2f)
                await client.send_message(message.channel, embed=embed)
                break

            if sheet["A" + str(i)].value == None:
                kk = 1
                break
            i = i + 1

        if kk == 1:
            if message.content.startswith('+unreal'):
                await client.send_message(message.channel,
                                          ":star: VIP버전 - 20유로 (2만 5,500원) 입니다.:asterisk:  Premium버전 - 15유로 (1만 9,100원) 입니다.")
            if message.content.startswith('+h4x'):
                await client.send_message(message.channel, ":star:VIP버전 - 12.99달러 (1만 4,607원) 입니다.")
            if message.content.startswith('+lolicon'):
                await client.send_message(message.channel,
                                          ":icecream:Ultimate버전 - 69.95유로 (9만 300원) 입니다.:icecream:Professional버전 - 49.95유로 (6만 4,400원) 입니다.:icecream:Lite 버전 - 19.95유로 (2만 5,700원) 입니다.")
            if message.content.startswith('+maximus'):
                await client.send_message(message.channel,
                                          ":star:MAX RECOVERY 버전 - 99.90헤알 (3만 200원) 입니다.:star:Premium 버전 - 75.90헤알 (2만 3,000원) 입니다.:star:VIP 버전 -  (1만 5,400원) 입니다.")
            if message.content.startswith('+maximus 정보'):
                await client.send_message(message.channel, "```1) 환불 불가능.\n2) 막시무스 계정 공유 금지 \n3) 업데이트는 무료로 진행됨 "
                                                           "\n4) MAX RECOVERY 버전은 라이센스가 3달로 제한되며, 3달마다 갱신할 수 있다."
                                                           "\n5) VIP, PREMIUM 버전은 라이센스가 1년으로 제한되며, 갱신할 수 있다. "
                                                           "\n6) VIP, PREMIUM 버전은 당신의 Social ID로 등록되며, "
                                                           "그 ID로 막시를 사용하는 한 얼마든지 다른 컴퓨터에서도 접속이 가능하다. "
                                                           "\n7) 막시무스는 GTA 온라인에서의 밴에 대한 어떤 책임도 지지 않는다.```")
            if message.content.startswith('+cubexis'):
                await client.send_message(message.channel, ":star:Platinum버전 - 29유로 (3만 7,460원) 입니다.")
            if message.content.startswith('+requiem'):
                await client.send_message(message.channel,
                                          ":icecream:레퀴엠 키 - 10달러 (약 1만 1,245원), :icecream:VIP 업그레이드 키 - 20달러 (2만 2,490원) 입니다.")
            if message.content.startswith('+diablo'):
                await client.send_message(message.channel, ":star:VIP버전 - 15유로 (약 19,132원) 입니다.")
            if message.content.startswith('+phantom-x'):
                await client.send_message(message.channel,
                                          ":icecream:Premium+ Edition - 1,990 루블 (약 33,392원), :icecream:Standard Edition - 990 루블 (약 16,612원) 입니다.")
            if message.content.startswith('+echo'):
                await client.send_message(message.channel, ":star:echo menu - 20달러 (약 2만 2600원) 입니다.")
            if message.content.startswith('+evolve'):
                await client.send_message(message.channel, "테이크 투 인터렉티브로 인해 막힘(구매 및 사용 불가)")
            if message.content.startswith('+maverick'):
                await client.send_message(message.channel,
                                          ":star:▪무료버전 : 무료 ▪유료버전 : 한 달에 5달러(약 5600원), 세 달에 10달러(약 11,100원), 평생버전 15달러(약 16,900원) ▪VIP평생버전 : 50달러(약 56,400원)")
            if message.content.startswith('+dynamic'):
                await client.send_message(message.channel,
                                          ":star:Premium 10.35파운드 (1만 5,035원) :star:VIP 15.40파운드 (2만 2,372원)")

            if message.content.startswith('+관리자'):
                if message.author.id == "249706579912294400":
                    state = message.content.split(" ")
                    file = openpyxl.load_workbook('상태.xlsx')
                    sheet = file.active
                    if sheet["A1"].value == str(state[1]):
                        await client.send_message(message.channel, "이미 말씀하신 상태에요.")
                    else:
                        sheet["A1"].value = str(state[1])
                        await client.send_message(message.channel, "현재 상태를 " + str(sheet["A1"].value) + "중으로 바꿨습니다.")
                        file.save('상태.xlsx')
                else:
                    msg = "권한이 없습니다."
                    embed = discord.Embed(description=msg, color=0xee2f2f)
                    await client.send_message(message.channel, embed=embed)

            if message.content.startswith('+메시지'):
                if message.author.id == "249706579912294400":
                    m = message.content[5:]
                    m = m.replace("<@249706579912294400>", "")
                    channel = client.get_channel("536203979428855818")
                    await client.send_message(channel, message.author.name + ": " + m)
                    await client.send_message(message.channel, "관리자방에 메시지를 보냈어요.")
                else:
                    msg = "권한이 없습니다."
                    embed = discord.Embed(description=msg, color=0xee2f2f)
                    await client.send_message(message.channel, embed=embed)

            if message.content.startswith('+추방'):
                msg = message.content.split(" ")
                role = ""
                for i in message.server.roles:
                    if i.id == "507513845225619456":
                        role = i
                        break
                msgg = msg[1][2:20]
                mem = message.server.get_member(msgg)
                if role in message.author.roles:
                    await client.kick(mem)
                    await client.send_message(message.channel, "회원을 추방했습니다.")
                else:
                    msg = "권한이 없습니다."
                    embed = discord.Embed(description=msg, color=0xee2f2f)
                    await client.send_message(message.channel, embed=embed)

            if message.content.startswith('+출근'):
                    msg1 = "출근 하셨습니다."
                    embed = discord.Embed(description=msg1, color=0xf30adc)
                    await client.send_message(message.channel, embed=embed)

            if message.content.startswith('+퇴근'):
                    msg1 = "퇴근 하셨습니다."
                    embed = discord.Embed(description=msg1, color=0xf30adc)
                    await client.send_message(message.channel, embed=embed)

            if message.content.startswith('+휴가'):
                    msg1 = "휴가중"
                    embed = discord.Embed(description=msg1, color=0xf30adc)
                    await client.send_message(message.channel, embed=embed)
            
            if message.content.startswith('+휴식'):
                    msg1 = "힐링중"
                    embed = discord.Embed(description=msg1, color=0xf30adc)
                    await client.send_message(message.channel, embed=embed)
            
            if message.content.startswith('+문의받음'):
                    msg1 = "다시 문의 받아요"
                    embed = discord.Embed(description=msg1, color=0xf30adc)
                    await client.send_message(message.channel, embed=embed)

            if message.content.startswith('+공지') and not message.content.startswith('+공지설정'):
                if message.author.id == "249706579912294400":
                    msg = message.content[23:]
                    content = message.content.split(" ")
                    channel = client.get_channel(content[1])
                    await client.send_message(channel, msg)
                    msg1 = "공지를 전송했습니다."
                    embed = discord.Embed(description=msg1, color=0x2ef079)
                    await client.send_message(message.channel, embed=embed)

                else:
                    msg = "권한이 없습니다."
                    embed = discord.Embed(description=msg, color=0xee2f2f)
                    await client.send_message(message.channel, embed=embed)

            if message.content.startswith("+전체공지"):
                if message.author.id == "249706579912294400":
                    content = message.content[6:]
                    file = openpyxl.load_workbook("가입서버.xlsx")
                    sheet = file.active
                    i = 1
                    while True:
                        if sheet["B" + str(i)].value != None:
                            channel = client.get_channel(sheet["B" + str(i)].value)
                            await client.send_message(channel, content)
                        else:
                            break
                        i += 1
                    msg = "전체공지를 전송했습니다."
                    embed = discord.Embed(description=msg, color=0x2ef079)
                    await client.send_message(message.channel, embed=embed)

                else:
                    msg = "권한이 없습니다."
                    embed = discord.Embed(description=msg, color=0xee2f2f)
                    await client.send_message(message.channel, embed=embed)

            if message.content.startswith("+플레이중"):
                if message.author.id == "249706579912294400":
                    await client.change_presence(game=discord.Game(name=str(message.content)[6:], type=1))
                else:
                    msg = "권한이 없습니다."
                    embed = discord.Embed(description=msg, color=0xee2f2f)
                    await client.send_message(message.channel, embed=embed)

            if message.content.startswith("+강화삭제"):
                if message.author.id == "249706579912294400":
                    msg = message.content.split(" ")
                    file = openpyxl.load_workbook("강화.xlsx")
                    sheet = file.active
                    i = 1
                    while True:
                        if sheet["A" + str(i)].value == msg[1]:
                            sheet["A" + str(i)].value = "-"
                            sheet["B" + str(i)].value = "-"
                            sheet["C" + str(i)].value = "-"
                            sheet["D" + str(i)].value = "-"
                            sheet["E" + str(i)].value = "-"
                            file.save('강화.xlsx')
                        if sheet["A" + str(i)].value == None:
                            break

                        i = i + 1
                    msg = "해당 아이디의 강화목록을 모두 삭제했습니다."
                    embed = discord.Embed(description=msg, color=0x2ef079)
                    await client.send_message(message.channel, embed=embed)

                else:
                    msg = "권한이 없습니다."
                    embed = discord.Embed(description=msg, color=0xee2f2f)
                    await client.send_message(message.channel, embed=embed)

            if message.content.startswith("+강화초기화"):
                if message.author.id == "249706579912294400":
                    file = openpyxl.load_workbook("강화.xlsx")
                    sheet = file.active
                    i = 1
                    while True:
                        if sheet["A" + str(i)].value == None:
                            break
                        if sheet["A" + str(i)].value != None:
                            sheet["A" + str(i)].value = None
                            sheet["B" + str(i)].value = None
                            sheet["C" + str(i)].value = None
                            sheet["D" + str(i)].value = None
                            sheet["E" + str(i)].value = None

                        i = i + 1
                    file.save('강화.xlsx')
                    msg = "강화목록을 초기화했습니다."
                    embed = discord.Embed(description=msg, color=0x2ef079)
                    await client.send_message(message.channel, embed=embed)

                else:
                    msg = "권한이 없습니다."
                    embed = discord.Embed(description=msg, color=0xee2f2f)
                    await client.send_message(message.channel, embed=embed)

            if message.content.startswith("+안내") and not message.content.startswith("+안내사항"):
                if message.author.id == "249706579912294400":
                    msg = message.content[4:]
                    f = open('안내.txt', 'w')
                    f.write(msg)
                    f.close()
                    msg = "안내 메시지를 등록했습니다."
                    embed = discord.Embed(description=msg, color=0x2ef079)
                    await client.send_message(message.channel, embed=embed)

                else:
                    msg = "권한이 없습니다."
                    embed = discord.Embed(description=msg, color=0xee2f2f)
                    await client.send_message(message.channel, embed=embed)

            if message.content.startswith('+차단'):
                if message.author.id == "249706579912294400":
                    msg = message.content.split(" ")
                    i = 1
                    while True:
                        if sheet1["A" + str(i)].value == msg[1]:
                            sheet1["A" + str(i)].value = "-"
                            file1.save("밴.xlsx")
                            msg = "차단을 해제했습니다."
                            embed = discord.Embed(description=msg, color=0x2ef079)
                            await client.send_message(message.channel, embed=embed)
                            break
                        if sheet1["A" + str(i)].value == None:
                            sheet1["A" + str(i)].value = str(msg[1])
                            file1.save("밴.xlsx")
                            msg = "이용자를 차단 했습니다."
                            embed = discord.Embed(description=msg, color=0x2ef079)
                            await client.send_message(message.channel, embed=embed)
                            break

                        i += 1
                else:
                    msg = "권한이 없습니다."
                    embed = discord.Embed(description=msg, color=0xee2f2f)
                    await client.send_message(message.channel, embed=embed)

            if message.content.startswith('+소개'):
                msg = "__**일반유저**__ 모두가 사용가능한 명령어" \
                      "\n`+도움말`" \
                      "\n`+일반명령어`" \
                      "\n`+음악명령어`" \
                      "\n`+게임명령어`" \
                      "\n\n__**서버소유자**__ 만 사용가능한 명령어" \
                      "\n`+공지설정 <채널명>`" \
                      "\n\n__**봇 소유자**__ 만 사용가능한 명령어" \
                      "\n`+관리자 <관리자상태>`" \
                      "\n`+플레이중 <내용>`" \
                      "\n`+공지 <채널id> <내용>`" \
                      "\n`+전체공지 <내용>`" \
                      "\n`+안내 <내용>`" \
                      "\n`+강화초기화`" \
                      "\n`+강화삭제 <유저id>`" \
                      "\n`+차단 <유저id>`" \
                      "\n`+서버삭제 <서버id>`"
                embed = discord.Embed(description=msg, color=0x2ef079)
                await client.send_message(message.channel, embed=embed)

            if message.content.startswith('+도움말'):
                msg = "`+소개` - 봇을 소개합니다." \
                      "\n`+일반명령어` - 일반 명령어" \
                      "\n`+음악명령어` - 음악기능 명령어" \
                      "\n`+게임명령어` - 강화게임 명령어" \
                      "\n`+서버소유자명령어` - 서버 소유자 전용 명령어"
                embed = discord.Embed(description=msg, color=0x2ef079)
                await client.send_message(message.channel, embed=embed)

            if message.content.startswith('+일반명령어'):
                msg = "**__일반 명령어__**" \
                      "\n`+안내사항` -> 공지나 안내메시지를 확인할수 있습니다." \
                      "\n`+회원가입` -> 봇에게 회원가입을 합니다." \
                      "\n`+가입회원` -> 가입한 회원수를 알려줍니다." \
                      "\n`+가입서버` -> 가입된 서버의 개수를 알려줍니다." \
                      "\n`+시간` -> 현재 시각을 알려줍니다." \
                      "\n`+검색 <내용>` -> 통합검색을 할 수 있습니다." \
                      "\n`+날씨 <지역>` -> 지역의 날씨를 말해줍니다." \
                      "\n`+초대` -> 초대링크를 받을수있습니다."
                embed = discord.Embed(description=msg, color=0x2ef079)
                await client.send_message(message.channel, embed=embed)

            if message.content.startswith('+게임명령어'):
                msg = "**__게임 명령어__**" \
                      "\n`+강화 <이름>` -> <이름>을 강화합니다." \
                      "\n`+강화랭킹` -> 현재서버 강화순위을 알려줍니다."
                embed = discord.Embed(description=msg, color=0x2ef079)
                await client.send_message(message.channel, embed=embed)

            if message.content.startswith('+음악명령어'):
                msg = "**__음악 명령어__**" \
                      "\n`+입장` -> 음성채널에 입장합니다." \
                      "\n`+재생 (URL 또는 곡명)` -> 노래를 재생합니다." \
                      "\n`+다음곡` -> 재생중인 노래를 넘기고 다음 노래를 재생합니다." \
                      "\n`+재생목록` -> 재생목록을 확인합니다." \
                      "\n`+중지` -> 노래를 일시정지 합니다." \
                      "\n`+중지해제` -> 일시정지한 노래를 다시 재생합니다." \
                      "\n`+끄기` -> 노래를 종료하고 음성채널을 나갑니다."
                embed = discord.Embed(description=msg, color=0x2ef079)
                await client.send_message(message.channel, embed=embed)

            if message.content.startswith('+서버소유자명령어'):
                msg = "**__서버 소유자 명령어__**" \
                      "\n`+공지설정 <채널명>` -> 공지가 전송되는 채널을 선택합니다."
                embed = discord.Embed(description=msg, color=0x2ef079)
                await client.send_message(message.channel, embed=embed)

            if message.content.startswith("+입장"):
                await client.send_message(message.channel, "보이스채널에 입장합니다.")

            if message.content.startswith("+끄기"):
                await client.send_message(message.channel, "보이스채널에서 퇴장합니다.")

            if message.content.startswith("+안내사항"):
                f = open("안내.txt")
                msg = f.read()
                f.close()
                embed = discord.Embed(description=msg, color=0x2ef079)
                await client.send_message(message.channel, embed=embed)

            if message.content.startswith("+검색"):
                ss = message.content[4:]
                ss = ss.split(" ")
                loc = ss[0]
                for i in range(1, len(ss)):
                    loc = loc + "+" + ss[i]
                search = message.content[7:] + "에 대한 검색결과입니다.\n\n[1. 지식백과](https://search.naver.com/search.naver?where=kdic&sm=tab_jum&query=" + loc + ")\n" + Knoledge.get_knoledge(loc)[:50] + "\n\n[2. 지식in](https://search.naver.com/search.naver?where=kin&sm=tab_jum&query=" + loc + ")\n" + AI.get_image(loc)[:50] + "\n\n[3. 웹사이트](https://search.naver.com/search.naver?where=webkr&sm=tab_jum&query=" + loc + ")\n" + Dictionary.get_deng(loc)[:50]
                embed = discord.Embed(description=search, color=0x2ef079)
                await client.send_message(message.channel, embed=embed)

            if message.content.startswith("+시간"):
                msg = str(datetime.datetime.today().year) + "년 " + str(datetime.datetime.today().month) + "월 " + str(datetime.datetime.today().day) + "일 " + str(datetime.datetime.today().hour) + "시 " + str(datetime.datetime.today().minute) + "분 " + str(datetime.datetime.today().second) + "초"
                embed = discord.Embed(description=msg, color=0x2ef079)
                await client.send_message(message.channel, embed=embed)

            if message.content.startswith("+날씨"):
                loc = message.content.split(" ")
                temp = Weather.get_weather(loc[1] + " 날씨")
                embed = discord.Embed(description=temp, color=0x2ef079)
                await client.send_message(message.channel, embed=embed)

            if message.content.startswith("+서버삭제") and message.author.id == "249706579912294400":
                channel = message.content.split(" ")
                file = openpyxl.load_workbook("가입서버.xlsx")
                sheet = file.active
                i = 1
                while True:
                    if sheet["A" + str(i)].value == str(message.server.id):
                        sheet["A" + str(i)].value = "-"
                        sheet["B" + str(i)].value = "-"
                        file.save('가입서버.xlsx')
                        break
                    if sheet["A" + str(i)].value == None:
                        break
                    i += 1
                leave = client.get_server(str(channel[1]))
                await client.leave_server(leave)

            if message.content.startswith("+회원가입"):
                file = openpyxl.load_workbook("가입회원.xlsx")
                sheet = file.active
                i = 1
                while True:
                    if sheet["A" + str(i)].value == str(message.author.id):
                        msg = "이미 가입된 회원입니다."
                        embed = discord.Embed(description=msg, color=0x2ef079)
                        await client.send_message(message.channel, embed=embed)
                        break

                    if sheet["A" + str(i)].value == None:
                        num = int(sheet["C1"].value) + 1
                        sheet["A" + str(i)].value = str(message.author.id)
                        sheet["C1"].value = int(sheet["C1"].value) + 1
                        msg = "**환영합니다! " + str(
                            num) + "번째 회원입니다.**"
                        embed = discord.Embed(description=msg, color=0x2ef079)
                        await client.send_message(message.channel, embed=embed)
                        file.save('가입회원.xlsx')
                        break

                    i += 1

            if message.content.startswith("+가입회원"):
                file = openpyxl.load_workbook("가입회원.xlsx")
                sheet = file.active
                msg = "현재 " + str(sheet["C1"].value) + "명이 회원가입을 했습니다."
                embed = discord.Embed(description=msg, color=0x2ef079)
                await client.send_message(message.channel, embed=embed)

            if message.content.startswith("+가입서버"):
                msg = "현재 가입한 서버는 " + str(len(client.servers)) + "개입니다."
                embed = discord.Embed(description=msg, color=0x2ef079)
                await client.send_message(message.channel, embed=embed)

            if message.content.startswith("+강화랭킹"):
                file = openpyxl.load_workbook("강화.xlsx")
                sheet = file.active
                list = []
                i = 1
                while True:
                    if sheet["E" + str(i)].value == str(message.server.id):
                        list.append("Lv.%03d `" % int(sheet["C" + str(i)].value) + str(sheet["B" + str(i)].value) + "` " + discord.utils.get(client.get_all_members(), id=str(sheet["A" + str(i)].value)).name)
                    if sheet["A" + str(i)].value == None:
                        break
                    i = i + 1
                a, b, c, d, e = "-", "-", "-", "-", "-"
                list.sort(reverse=True)
                if len(list) == 0:
                    a, b, c, d, e = "-", "-", "-", "-", "-"
                if len(list) == 1:
                    a, b, c, d, e = list[0], "-", "-", "-", "-"
                if len(list) == 2:
                    a, b, c, d, e = list[0], list[1], "-", "-", "-"
                if len(list) == 3:
                    a, b, c, d, e = list[0], list[1], list[2], "-", "-"
                if len(list) == 4:
                    a, b, c, d, e = list[0], list[1], list[2], list[3], "-"
                if len(list) == 5:
                    a, b, c, d, e = list[0], list[1], list[2], list[3], list[4]
                mm = "1. " + a + "\n2. " + b + "\n3. " + c + "\n4. " + d + "\n5. " + e
                embed = discord.Embed(description=mm)
                await client.send_message(message.channel, embed=embed)

            if message.content.startswith("+강화") and not message.content.startswith("+강화랭킹") and not message.content.startswith("+강화삭제") and not message.content.startswith("+강화초기화"):
                file = openpyxl.load_workbook("강화.xlsx")
                sheet = file.active
                item = message.content.split(" ")
                num = 0
                i = 1
                while True:
                    if sheet["A" + str(i)].value == message.author.id:
                        num = num + 1
                        if sheet["B" + str(i)].value == item[1]:
                            if int(sheet["D" + str(i)].value) <= int(datetime.datetime.today().strftime("%Y%m%d%H%M%S")):
                                time = datetime.datetime.today() + datetime.timedelta(seconds=30)
                                sheet["D" + str(i)].value = time.strftime("%Y%m%d%H%M%S")
                                level = int(sheet["C" + str(i)].value)
                                rank = enforce.enf(level)[1]
                                a = random.randint(1, 100)
                                if a <= enforce.enf(level)[0]:
                                    up = random.randint(1, 30)
                                    level1 = level + up
                                    rank1 = enforce.enf(level + up)[1]
                                    sheet["C" + str(i)].value = level1
                                    mm = "강화에 성공하였습니다.\n`%s` Lv.%s -> Lv.%s `%s` -> `%s`" % (sheet["B" + str(i)].value, str(level), str(level1), rank, rank1)
                                    embed = discord.Embed(description=mm, color=0x2037bb)
                                    await client.send_message(message.channel, embed=embed)

                                elif a > 95:
                                    sheet["C" + str(i)].value = "0"
                                    mm = "`%s`가 터졌습니다.\nLv.0 `알`" % sheet["B" + str(i)].value
                                    embed = discord.Embed(description=mm, color=0xee2f2f)
                                    await client.send_message(message.channel, embed=embed)

                                else:
                                    down = random.randint(1, 25)
                                    level1 = level - down
                                    rank1 = enforce.enf(level - down)[1]
                                    sheet["C" + str(i)].value = level1
                                    mm = "강화에 실패하였습니다.\n`%s` Lv.%s -> Lv.%s `%s` -> `%s`" % (sheet["B" + str(i)].value, str(level), str(level1), rank, rank1)
                                    embed = discord.Embed(description=mm, color=0xee2f2f)
                                    await client.send_message(message.channel, embed=embed)

                                file.save('강화.xlsx')
                                break

                            else:
                                mm = "30초 쿨타임이 지나지 않았습니다."
                                embed = discord.Embed(description=mm, color=0xee2f2f)
                                await client.send_message(message.channel, embed=embed)
                                break

                    if sheet["A" + str(i)].value == None:
                        if num == 2:
                            msg = "강화 개수 제한이 꽉 찼습니다."
                            embed = discord.Embed(description=msg, color=0xee2f2f)
                            await client.send_message(message.channel, embed=embed)
                        else:
                            sheet["A" + str(i)].value = str(message.author.id)
                            sheet["B" + str(i)].value = item[1]
                            sheet["C" + str(i)].value = 0
                            sheet["D" + str(i)].value = 0
                            sheet["E" + str(i)].value = str(message.server.id)
                            mm = "`%s`를 새로 만들었습니다." % item[1]
                            embed = discord.Embed(description=mm, color=0x2ef079)
                            await client.send_message(message.channel, embed=embed)
                        file.save('강화.xlsx')
                        break

                    i = i + 1

            if message.content.startswith("+공지설정"):
                if message.author == message.server.owner:
                    file = openpyxl.load_workbook("가입서버.xlsx")
                    sheet = file.active
                    msg = message.content.split(" ")
                    channel = discord.utils.get(message.server.channels, name=msg[1])
                    i = 1
                    while True:
                        if sheet["A" + str(i)].value == str(message.server.id) or sheet["A" + str(i)].value == None:
                            sheet["A" + str(i)].value = str(message.server.id)
                            sheet["B" + str(i)].value = str(channel.id)
                            msg = "공지채널을 등록했습니다."
                            embed = discord.Embed(description=msg, color=0x2ef079)
                            await client.send_message(message.channel, embed=embed)
                            file.save('가입서버.xlsx')
                            break
                        i += 1

                else:
                    msg = "권한이 없습니다."
                    embed = discord.Embed(description=msg, color=0xee2f2f)
                    await client.send_message(message.channel, embed=embed)


client.run('NTM2MTkwNTk4ODA2NTAzNDM1.DyTdPA.Wj4c8CFJgyMsan85WQWGcKpkCa0')




